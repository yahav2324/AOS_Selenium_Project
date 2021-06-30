from openpyxl import load_workbook
workbook = load_workbook(filename=r"C:\Users\yahav\Desktop\AOS test.xlsx")
sheet = workbook.active
def data(a, b):
    tests = [[] for i in range(1, 11)]
    i = 0
    for col in sheet.iter_cols(min_col=3, max_col=12, min_row=2, max_row=25, values_only=True):
        for field in col:
            if field is not None:
                tests[i].append(field)
        i += 1
    return tests[a][b]


